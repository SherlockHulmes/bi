import json
import re
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.utils import timezone
from sqlalchemy import text

from .models import LineageScan, LineageTable, LineageEdge
from .sql_parser import parse_event_sql, get_table_columns, get_table_comment
from sql_scripts.models import DbConnection
from core.utils import get_engine_from_db_conn
from core.decorators import admin_required

logger = logging.getLogger('data_lineage')


@login_required
def scan_list(request):
    """扫描记录列表"""
    scans = LineageScan.objects.select_related('source_db', 'scanned_by').all()
    db_connections = DbConnection.objects.filter(is_active=True)
    return render(request, 'data_lineage/scans.html', {
        'scans': scans,
        'db_connections': db_connections,
    })


def _do_scan(db_conn, user, existing_scan=None):
    """执行扫描核心逻辑，返回 (scan, debug_info)"""
    engine = get_engine_from_db_conn(db_conn)

    with engine.connect() as conn:
        result = conn.execute(text(
            "SELECT EVENT_NAME, EVENT_DEFINITION, EVENT_TYPE, EXECUTE_AT, "
            "INTERVAL_VALUE, INTERVAL_FIELD, STATUS "
            "FROM information_schema.EVENTS "
            "WHERE EVENT_SCHEMA = :db"
        ), {'db': db_conn.database})
        events = []
        for row in result:
            event_name = row[0]
            event_sql = row[1] or ''
            logger.info(f"[扫描] 事件: {event_name}, SQL前100字符: {event_sql[:100]}")
            events.append({'name': event_name, 'sql': event_sql})
    logger.info(f"[扫描] 共读取到 {len(events)} 个事件")

    if existing_scan:
        scan = existing_scan
        scan.tables.all().delete()
        scan.edges.all().delete()
        scan.event_count = len(events)
        scan.table_count = 0
        scan.edge_count = 0
        scan.save()
    else:
        scan = LineageScan(source_db=db_conn, scanned_by=user, event_count=len(events))
        scan.save()

    all_tables = {}
    all_edges = []
    debug_info = []
    exclude_words = {'SELECT', 'DUAL', 'INFORMATION_SCHEMA', 'SET', 'VALUES', 'ON', 'AS'}

    # 第一轮：收集所有已知数据库名（从 information_schema + SQL 中提取）
    known_db_names = {db_conn.database}

    # 从 information_schema 获取所有数据库名，防止数据库名被误识别为表名
    try:
        with engine.connect() as conn:
            schema_result = conn.execute(text("SELECT SCHEMA_NAME FROM information_schema.SCHEMATA"))
            for row in schema_result:
                known_db_names.add(row[0])
    except Exception:
        logger.warning("[扫描] 无法查询 information_schema.SCHEMATA，跳过")

    # 也从 SQL 中 db.table 格式提取数据库名（覆盖 information_schema 中没有的库）
    db_pattern = r'(?:FROM|JOIN|INTO)\s+`?(\w+)`?\.`?\w+`?'
    for event in events:
        for m in re.finditer(db_pattern, event['sql'], re.IGNORECASE):
            db_n = m.group(1)
            if db_n.upper() not in exclude_words:
                known_db_names.add(db_n)
    logger.info(f"[扫描] 已识别的数据库名: {known_db_names}")

    for event in events:
        event_name = event['name']
        event_sql = event['sql']
        parsed = parse_event_sql(event_sql, known_db_names=known_db_names)

        # 安全过滤：排除已知数据库名和 CTE 名被误识别为目标表/源表
        known_dbs_upper = {n.upper() for n in known_db_names}
        # 从 parser 结果中提取 CTE 名（parser 内部已排除，但 views 也需要排除）
        cte_upper = set()
        for _m in re.finditer(r'(?:WITH|,)\s*(\w+)\s+AS\s*\(', event_sql, re.IGNORECASE):
            cte_upper.add(_m.group(1).upper())
        
        exclude_upper = known_dbs_upper | cte_upper
        target_tables = [t for t in parsed['target_tables'] if t.upper() not in exclude_upper]
        # 同样过滤源表
        source_tables = []
        for s in parsed['source_tables']:
            if '.' in s:
                s_db, s_name = s.split('.', 1)
                if s_db.upper() not in exclude_words and s_name.upper() not in exclude_upper:
                    source_tables.append(s)
            else:
                if s.upper() not in exclude_upper:
                    source_tables.append(s)

        logger.info(f"[解析] 事件 {event_name}: 目标表={target_tables}, 源表={source_tables}")
        debug_info.append({'name': event_name, 'sql_short': event_sql[:200], 'target_tables': target_tables, 'source_tables': source_tables})

        for t_table in target_tables:
            table_key = f"{db_conn.database}.{t_table}"
            if table_key not in all_tables:
                comment = get_table_comment(engine, db_conn.database, t_table)
                columns = get_table_columns(engine, db_conn.database, t_table)
                lt = LineageTable(scan=scan, table_name=t_table, database_name=db_conn.database, table_comment=comment, columns_json=json.dumps(columns, ensure_ascii=False), is_source=False)
                lt.save()
                all_tables[table_key] = lt

        for s_table in source_tables:
            if '.' in s_table:
                s_db, s_name = s_table.split('.', 1)
            else:
                s_db, s_name = db_conn.database, s_table
            if s_db.upper() in exclude_words:
                continue
            table_key = f"{s_db}.{s_name}"
            if table_key not in all_tables:
                comment = get_table_comment(engine, s_db, s_name)
                columns = get_table_columns(engine, s_db, s_name)
                lt = LineageTable(scan=scan, table_name=s_name, database_name=s_db, table_comment=comment, columns_json=json.dumps(columns, ensure_ascii=False), is_source=True)
                lt.save()
                all_tables[table_key] = lt

        for t_table in target_tables:
            for s_table in source_tables:
                if '.' in s_table:
                    s_db, s_name = s_table.split('.', 1)
                else:
                    s_db, s_name = db_conn.database, s_table
                if s_db.upper() in exclude_words:
                    continue
                edge = LineageEdge(scan=scan, source_table=s_name, source_database=s_db, target_table=t_table, target_database=db_conn.database, event_name=event_name, event_sql=event_sql)
                edge.save()
                all_edges.append(edge)

    scan.table_count = len(all_tables)
    scan.edge_count = len(all_edges)
    scan.save()
    return scan, debug_info, len(events), len(all_tables), len(all_edges)


@login_required
@require_POST
def scan_execute(request):
    """执行扫描：读取 EVENTS → 解析血缘 → 存储结果"""
    db_id = request.POST.get('source_db')
    if not db_id:
        messages.error(request, '请选择数据库连接')
        return redirect('data_lineage:scan_list')

    try:
        db_conn = DbConnection.objects.get(pk=db_id, is_active=True)
    except DbConnection.DoesNotExist:
        messages.error(request, '数据库连接不存在')
        return redirect('data_lineage:scan_list')

    try:
        scan, debug_info, event_count, table_count, edge_count = _do_scan(db_conn, request.user)
        messages.success(request, f'扫描完成：{event_count} 个事件，{table_count} 张表，{edge_count} 条依赖关系')
        request.session['scan_debug'] = debug_info
        return redirect('data_lineage:graph', scan_id=scan.pk)
    except Exception as e:
        logger.error(f"扫描失败: {e}")
        messages.error(request, f'扫描失败: {str(e)[:200]}')
        return redirect('data_lineage:scan_list')


@login_required
@require_POST
def scan_update(request, scan_id):
    """更新已有扫描记录（重新扫描）"""
    scan = get_object_or_404(LineageScan, pk=scan_id)
    try:
        scan, debug_info, event_count, table_count, edge_count = _do_scan(scan.source_db, request.user, existing_scan=scan)
        messages.success(request, f'更新完成：{event_count} 个事件，{table_count} 张表，{edge_count} 条依赖关系')
        request.session['scan_debug'] = debug_info
        return redirect('data_lineage:graph', scan_id=scan.pk)
    except Exception as e:
        logger.error(f"更新失败: {e}")
        messages.error(request, f'更新失败: {str(e)[:200]}')
        return redirect('data_lineage:scan_list')


@login_required
def graph_view(request, scan_id):
    """血缘图页面"""
    scan = get_object_or_404(LineageScan, pk=scan_id)
    tables = list(scan.tables.all().values('id', 'table_name', 'database_name', 'table_comment', 'is_source'))
    edges = list(scan.edges.all().values('source_table', 'source_database', 'target_table', 'target_database', 'event_name'))

    # 构建 vis.js 数据 - 每个中间表独立显示其基础表
    # 按目标表分组边，每个目标表有自己的基础表节点
    source_db_names = {t['database_name'] for t in tables if t['is_source']}
    target_tables_set = {t['table_name'] for t in tables if not t['is_source']}

    # 收集所有有效的边（基础表→中间表）
    valid_edges = []
    for e in edges:
        if e['source_database'] not in source_db_names:
            continue
        valid_edges.append(e)

    # 按目标表分组
    target_groups = {}
    for e in valid_edges:
        tgt = e['target_table']
        if tgt not in target_groups:
            target_groups[tgt] = []
        target_groups[tgt].append(e)

    nodes = []
    vis_edges = []
    used_targets = set()

    # 布局参数：每组纵向间距、子表相对父表纵向偏移、子表横向间距
    GROUP_HEIGHT = 320        # 每组占用的纵向总高度（含留白）
    CHILD_Y_OFFSET = 140      # 子表相对父表的纵向偏移
    CHILD_SPACING = 200       # 子表之间的横向间距
    GROUP_PADDING = 300       # 组间额外横向间距（用于错开不同组避免左右重叠）

    group_index = 0

    for tgt_table, group_edges in target_groups.items():
        tgt_comment = ''
        for t in tables:
            if t['table_name'] == tgt_table and not t['is_source']:
                tgt_comment = t['table_comment'] or ''
                break
        tgt_label = tgt_table
        if tgt_comment:
            tgt_label += f"\n{tgt_comment[:20]}"

        # 计算子表数量，用于居中布局
        num_children = len(group_edges)
        # 子表总宽度
        children_total_width = (num_children - 1) * CHILD_SPACING if num_children > 1 else 0
        # 子表起始 x（居中对齐，父节点在子节点组的中心）
        children_start_x = -children_total_width // 2

        # 父节点 x 取子节点组的中心，y 按组索引递增
        parent_x = 0
        parent_y = group_index * GROUP_HEIGHT

        # 中间表节点（父节点）
        if tgt_table not in used_targets:
            nodes.append({
                'id': tgt_table,
                'label': tgt_label,
                'title': f"{group_edges[0]['target_database']}.{tgt_table}\n{tgt_comment}",
                'table_name': tgt_table,
                'color': '#2196F3',
                'shape': 'box',
                'font': {'color': 'white', 'size': 14},
                'db_name': group_edges[0]['target_database'],
                'table_comment': tgt_comment,
                'is_source': False,
                'level': 0,
                'x': parent_x,
                'y': parent_y,
                'fixed': True,
            })
            used_targets.add(tgt_table)

        # 基础表节点（子节点），横向排列在父节点下方
        child_y = parent_y + CHILD_Y_OFFSET
        for idx, e in enumerate(group_edges):
            src_unique_id = f"{tgt_table}__{e['source_table']}_{idx}"
            src_label = e['source_table']
            src_comment = ''
            for t in tables:
                if t['table_name'] == e['source_table'] and t['database_name'] == e['source_database']:
                    src_comment = t['table_comment'] or ''
                    break
            if src_comment:
                src_label += f"\n{src_comment[:20]}"

            child_x = children_start_x + idx * CHILD_SPACING

            nodes.append({
                'id': src_unique_id,
                'label': src_label,
                'title': f"{e['source_database']}.{e['source_table']}\n{src_comment}",
                'table_name': e['source_table'],
                'color': '#4CAF50',
                'shape': 'box',
                'font': {'color': 'white', 'size': 12},
                'db_name': e['source_database'],
                'table_comment': src_comment,
                'is_source': True,
                'level': 1,
                'x': child_x,
                'y': child_y,
                'fixed': True,
            })

            vis_edges.append({
                'from': src_unique_id,
                'to': tgt_table,
                'title': e['event_name'],
                'label': '',
                'arrows': 'to',
                'font': {'size': 10, 'align': 'middle'},
            })

        group_index += 1

    # 获取调试信息
    debug_info = request.session.pop('scan_debug', None)

    return render(request, 'data_lineage/graph.html', {
        'scan': scan,
        'nodes_json': json.dumps(nodes, ensure_ascii=False),
        'edges_json': json.dumps(vis_edges, ensure_ascii=False),
        'tables_json': json.dumps(tables, ensure_ascii=False),
        'debug_info': debug_info,
    })


@login_required
def table_detail(request, scan_id):
    """表详情 API"""
    db_name = request.GET.get('db', '')
    table_name = request.GET.get('table', '')

    scan = get_object_or_404(LineageScan, pk=scan_id)
    table_obj = scan.tables.filter(database_name=db_name, table_name=table_name).first()

    if not table_obj:
        return JsonResponse({'error': '表不存在'}, status=404)

    return JsonResponse({
        'table_name': table_obj.table_name,
        'database_name': table_obj.database_name,
        'table_comment': table_obj.table_comment,
        'columns': table_obj.get_columns(),
        'is_source': table_obj.is_source,
    })


@login_required
def export_excel(request, scan_id):
    """导出血缘数据为 Excel"""
    import io
    import openpyxl
    from django.utils.encoding import smart_str

    scan = get_object_or_404(LineageScan, pk=scan_id)

    wb = openpyxl.Workbook()

    # Sheet 1: 表清单
    ws1 = wb.active
    ws1.title = "表清单"
    ws1.append(['表名', '数据库', '类型', '注释'])
    for t in scan.tables.all():
        ws1.append([
            t.table_name,
            t.database_name,
            '基础表' if t.is_source else '中间表',
            t.table_comment,
        ])

    # Sheet 2: 字段明细
    ws2 = wb.create_sheet("字段明细")
    ws2.append(['表名', '数据库', '字段名', '字段类型', '注释'])
    for t in scan.tables.all():
        for col in t.get_columns():
            ws2.append([
                t.table_name,
                t.database_name,
                col.get('name', ''),
                col.get('type', ''),
                col.get('comment', ''),
            ])

    # Sheet 3: 依赖关系
    ws3 = wb.create_sheet("依赖关系")
    ws3.append(['源表', '源数据库', '目标表', '目标数据库', '事件名'])
    for e in scan.edges.all():
        ws3.append([
            e.source_table,
            e.source_database,
            e.target_table,
            e.target_database,
            e.event_name,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"数据血缘_{scan.source_db.name}_{timezone.localtime(scan.scanned_at).strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{smart_str(filename)}"'
    return response


@login_required
@admin_required
@require_POST
def scan_delete(request, scan_id):
    """删除扫描记录（需要管理员权限）"""
    scan = get_object_or_404(LineageScan, pk=scan_id)
    scan.delete()
    messages.success(request, '扫描记录已删除')
    return redirect('data_lineage:scan_list')
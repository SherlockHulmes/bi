import time
import logging
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from sqlalchemy import text

from .models import SqlScript, ExecutionLog, DbConnection
from core.decorators import admin_required
from core.utils import get_engine_from_db_conn, get_default_engine

logger = logging.getLogger('sql_scripts')


@login_required
def script_list(request):
    """脚本列表"""
    scripts = SqlScript.objects.select_related('target_db').all()
    return render(request, 'sql_scripts/list.html', {'scripts': scripts})


@login_required
def script_create(request):
    """创建脚本"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '')
        content = request.POST.get('content', '')
        timeout = int(request.POST.get('timeout', 300))
        target_db_id = request.POST.get('target_db')
        uploaded_file = request.FILES.get('file')

        if not name:
            messages.error(request, '脚本名称不能为空')
            return render(request, 'sql_scripts/create.html', {
                'db_connections': DbConnection.objects.filter(is_active=True)
            })

        script = SqlScript(
            name=name,
            description=description,
            content=content,
            timeout=timeout,
            created_by=request.user,
        )
        if target_db_id:
            try:
                script.target_db = DbConnection.objects.get(pk=target_db_id, is_active=True)
            except DbConnection.DoesNotExist:
                pass

        if uploaded_file:
            script.file = uploaded_file
            if not content:
                try:
                    content_bytes = uploaded_file.read()
                    script.content = content_bytes.decode('utf-8')
                except Exception:
                    pass
                uploaded_file.seek(0)
        script.save()
        messages.success(request, f'脚本 "{name}" 已创建')
        return redirect('scripts:detail', pk=script.pk)

    return render(request, 'sql_scripts/create.html', {
        'db_connections': DbConnection.objects.filter(is_active=True)
    })


@login_required
def script_detail(request, pk):
    """脚本详情"""
    script = get_object_or_404(SqlScript, pk=pk)
    logs = script.logs.all()[:20]
    sql_content = script.get_sql_content()
    return render(request, 'sql_scripts/detail.html', {
        'script': script,
        'logs': logs,
        'sql_content': sql_content,
    })


@login_required
@require_POST
def script_execute(request, pk):
    """手动执行 SQL 脚本"""
    script = get_object_or_404(SqlScript, pk=pk, is_active=True)
    sql_content = script.get_sql_content()

    if not sql_content or not sql_content.strip():
        return JsonResponse({'success': False, 'error': '脚本内容为空'})

    # 创建执行日志
    log = ExecutionLog.objects.create(
        script=script,
        trigger_type='manual',
        status='running',
        executed_by=request.user,
    )

    try:
        if script.target_db:
            engine = get_engine_from_db_conn(script.target_db)
        else:
            engine = get_default_engine()

        from core.utils import execute_sql_statements
        all_results, all_output, total_affected, duration = execute_sql_statements(engine, sql_content, max_rows=200)

        # 为 UI 添加 truncated 标记
        for r in all_results:
            if 'total_rows' in r:
                r['truncated'] = r['total_rows'] > 200

        log.status = 'success'
        log.output = json.dumps(all_results, ensure_ascii=False, default=str)
        log.duration = round(duration, 2)
        log.row_count = total_affected
        log.save()

        logger.info(f"脚本执行成功: {script.name}, 耗时 {duration:.2f}s by {request.user}")
        return JsonResponse({
            'success': True,
            'message': f'执行成功，耗时 {duration:.2f} 秒',
            'results': all_results,
            'duration': log.duration,
            'row_count': total_affected,
        })
    except Exception as e:
        log.status = 'failed'
        log.error_message = str(e)
        log.save()

        logger.error(f"脚本执行失败: {script.name}, 错误: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def script_edit(request, pk):
    """编辑脚本"""
    script = get_object_or_404(SqlScript, pk=pk)
    sql_content = script.get_sql_content()

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '')
        content = request.POST.get('content', '')
        timeout = int(request.POST.get('timeout', 300))
        target_db_id = request.POST.get('target_db')
        is_active = request.POST.get('is_active') == 'on'
        uploaded_file = request.FILES.get('file')

        if not name:
            messages.error(request, '脚本名称不能为空')
            return redirect('scripts:edit', pk=pk)

        script.name = name
        script.description = description
        script.content = content
        script.timeout = timeout
        script.is_active = is_active

        if target_db_id:
            try:
                script.target_db = DbConnection.objects.get(pk=target_db_id, is_active=True)
            except DbConnection.DoesNotExist:
                script.target_db = None
        else:
            script.target_db = None

        if uploaded_file:
            script.file = uploaded_file
            if not content:
                try:
                    content_bytes = uploaded_file.read()
                    script.content = content_bytes.decode('utf-8')
                except Exception:
                    pass
                uploaded_file.seek(0)

        script.save()
        messages.success(request, f'脚本 "{name}" 已更新')
        return redirect('scripts:detail', pk=pk)

    return render(request, 'sql_scripts/edit.html', {
        'script': script,
        'sql_content': sql_content,
        'db_connections': DbConnection.objects.filter(is_active=True),
    })


@login_required
def script_download_result(request, pk):
    """下载执行结果为 Excel 文件"""
    import io
    import json
    from django.http import HttpResponse
    from django.utils import timezone
    import openpyxl

    log_id = request.GET.get('log_id')
    if log_id:
        try:
            log = ExecutionLog.objects.get(pk=log_id, script_id=pk)
        except ExecutionLog.DoesNotExist:
            messages.error(request, '日志不存在')
            return redirect('scripts:detail', pk=pk)
    else:
        log = ExecutionLog.objects.filter(script_id=pk, status='success').order_by('-executed_at').first()
        if not log:
            messages.error(request, '没有可下载的执行结果')
            return redirect('scripts:detail', pk=pk)

    if not log.output:
        messages.error(request, '执行结果为空')
        return redirect('scripts:detail', pk=pk)

    try:
        results = json.loads(log.output)
    except json.JSONDecodeError:
        # 旧格式纯文本，直接返回文本文件
        response = HttpResponse(log.output, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="result_{pk}_{log.pk}.txt"'
        return response

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for idx, result in enumerate(results):
        if 'columns' in result and 'rows' in result:
            ws = wb.create_sheet(title=f"结果集{idx + 1}")
            ws.append(result['columns'])
            for row in result['rows']:
                ws.append([str(v) if v is not None else '' for v in row])
        elif 'message' in result:
            ws = wb.create_sheet(title=f"信息{idx + 1}")
            ws.append([result['message']])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    script = get_object_or_404(SqlScript, pk=pk)
    filename = f"{script.name}_{log.executed_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@admin_required
def script_delete(request, pk):
    """删除脚本（需要管理员权限）"""
    script = get_object_or_404(SqlScript, pk=pk)
    if request.method == 'POST':
        name = script.name
        script.delete()
        messages.success(request, f'脚本 "{name}" 已删除')
        return redirect('scripts:list')
    return render(request, 'sql_scripts/confirm_delete.html', {'script': script})
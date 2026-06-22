import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.utils import timezone

from .models import ScheduledTask, ScheduledTaskLog
from sql_scripts.models import SqlScript, ExecutionLog
from notifications.base import send_notification, format_task_notification, format_status_notification
from core.decorators import admin_required

logger = logging.getLogger('scheduler')


@login_required
def task_list(request):
    """任务列表"""
    tasks = ScheduledTask.objects.select_related('script').all()
    return render(request, 'scheduler/tasks.html', {'tasks': tasks})


@login_required
def task_create(request):
    """创建定时任务"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        script_id = request.POST.get('script_id')
        schedule_type = request.POST.get('schedule_type', 'daily')
        hour = int(request.POST.get('hour', 8))
        minute = int(request.POST.get('minute', 0))
        notify_channel = request.POST.get('notify_channel', 'none')
        days = request.POST.get('days', '')
        day_of_month = int(request.POST.get('day_of_month', 1))
        cron_expr = request.POST.get('cron_expr', '').strip()

        if not name or not script_id:
            messages.error(request, '请填写完整信息（任务名称和关联脚本必填）')
            scripts = SqlScript.objects.filter(is_active=True)
            return render(request, 'scheduler/create.html', {'scripts': scripts})

        if schedule_type == 'weekly' and not days:
            messages.error(request, '每周调度需要选择星期几')
            scripts = SqlScript.objects.filter(is_active=True)
            return render(request, 'scheduler/create.html', {'scripts': scripts})

        script = get_object_or_404(SqlScript, pk=script_id, is_active=True)
        email_recipients = request.POST.get('email_recipients', '').strip()
        task = ScheduledTask(
            name=name,
            script=script,
            schedule_type=schedule_type,
            schedule_hour=hour,
            schedule_minute=minute,
            schedule_days=days,
            schedule_day_of_month=day_of_month,
            cron_expr=cron_expr if schedule_type == 'custom' else '',
            notify_channel=notify_channel,
            email_recipients=email_recipients,
            created_by=request.user,
        )
        task.save()
        messages.success(request, f'定时任务 "{name}" 已创建')
        return redirect('scheduler:list')

    scripts = SqlScript.objects.filter(is_active=True)
    return render(request, 'scheduler/create.html', {'scripts': scripts})


@login_required
def task_edit(request, pk):
    """编辑定时任务"""
    task = get_object_or_404(ScheduledTask, pk=pk)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        script_id = request.POST.get('script_id')
        schedule_type = request.POST.get('schedule_type', 'daily')
        hour = int(request.POST.get('hour', 8))
        minute = int(request.POST.get('minute', 0))
        notify_channel = request.POST.get('notify_channel', 'none')
        days = request.POST.get('days', '')
        day_of_month = int(request.POST.get('day_of_month', 1))
        cron_expr = request.POST.get('cron_expr', '').strip()
        is_enabled = request.POST.get('is_enabled') == 'on'

        if not name or not script_id:
            messages.error(request, '请填写完整信息')
            scripts = SqlScript.objects.filter(is_active=True)
            return render(request, 'scheduler/edit.html', {'task': task, 'scripts': scripts})

        script = get_object_or_404(SqlScript, pk=script_id, is_active=True)
        task.name = name
        task.script = script
        task.schedule_type = schedule_type
        task.schedule_hour = hour
        task.schedule_minute = minute
        task.schedule_days = days
        task.schedule_day_of_month = day_of_month
        task.cron_expr = cron_expr if schedule_type == 'custom' else ''
        task.email_recipients = request.POST.get('email_recipients', '').strip()
        task.notify_channel = notify_channel
        task.is_enabled = is_enabled
        task.save()

        messages.success(request, f'定时任务 "{name}" 已更新')
        return redirect('scheduler:list')

    scripts = SqlScript.objects.filter(is_active=True)
    return render(request, 'scheduler/edit.html', {'task': task, 'scripts': scripts})


@login_required
@require_POST
def task_copy(request, pk):
    """复制定时任务"""
    original = get_object_or_404(ScheduledTask, pk=pk)
    new_task = ScheduledTask(
        name=f"{original.name} - 副本",
        script=original.script,
        schedule_type=original.schedule_type,
        schedule_hour=original.schedule_hour,
        schedule_minute=original.schedule_minute,
        schedule_days=original.schedule_days,
        schedule_day_of_month=original.schedule_day_of_month,
        cron_expr=original.cron_expr,
        notify_channel=original.notify_channel,
        email_recipients=original.email_recipients,
        is_enabled=False,
        created_by=request.user,
    )
    new_task.save()
    messages.success(request, f'已复制任务 "{original.name}" → "{new_task.name}"')
    return redirect('scheduler:edit', pk=new_task.pk)


@login_required
@require_POST
def task_toggle(request, pk):
    """启停任务"""
    task = get_object_or_404(ScheduledTask, pk=pk)
    task.is_enabled = not task.is_enabled
    task.save()
    status = '启用' if task.is_enabled else '停用'
    messages.success(request, f'任务 "{task.name}" 已{status}')
    return redirect('scheduler:list')


@login_required
@admin_required
@require_POST
def task_delete(request, pk):
    """删除任务（需要管理员权限）"""
    task = get_object_or_404(ScheduledTask, pk=pk)
    name = task.name
    task.delete()
    messages.success(request, f'任务 "{name}" 已删除')
    return redirect('scheduler:list')


@login_required
@require_POST
def task_notify(request, pk):
    """手动发送通知（不执行 SQL）"""
    task = get_object_or_404(ScheduledTask, pk=pk)
    if task.notify_channel == 'none':
        messages.warning(request, f'任务 "{task.name}" 未配置通知渠道')
        return redirect('scheduler:list')

    try:
        content = format_status_notification(
            task_name=task.name,
            schedule_display=task.get_schedule_display(),
            is_enabled=task.is_enabled,
            last_run_at=task.last_run_at,
            last_status=task.last_status,
        )
        send_notification(
            channel=task.notify_channel,
            title='任务状态通知',
            content=content
        )
        messages.success(request, f'✅ 通知已发送到 {task.get_notify_channel_display()}')
    except Exception as e:
        messages.error(request, f'❌ 通知发送失败: {str(e)[:200]}')

    return redirect('scheduler:list')


@login_required
@require_POST
def task_execute(request, pk):
    """手动执行定时任务（同步，不需要 Celery）"""
    from core.utils import get_engine_from_db_conn, get_default_engine, execute_sql_statements

    task = get_object_or_404(ScheduledTask, pk=pk)
    script = task.script
    sql_content = script.get_sql_content()

    if not sql_content or not sql_content.strip():
        messages.error(request, '脚本内容为空，无法执行')
        return redirect('scheduler:list')

    task_log = ScheduledTaskLog.objects.create(
        task=task,
        task_name=task.name,
        script_name=script.name,
        sql_executed=sql_content,
        trigger_type='manual',
        status='success',
        notify_channel=task.notify_channel if task.notify_channel != 'none' else '',
        notify_target=task.email_recipients or '',
        executed_by=request.user,
    )

    log = ExecutionLog.objects.create(
        script=script,
        trigger_type='manual',
        status='running',
        executed_by=request.user,
    )

    try:
        if script.target_db:
            engine = get_engine_from_db_conn(script.target_db)
        else:
            engine = get_default_engine()

        all_results, all_output, total_affected, duration = execute_sql_statements(engine, sql_content)

        log.status = 'success'
        log.output = '\n\n---\n\n'.join(all_output)
        log.duration = round(duration, 2)
        log.row_count = total_affected
        log.save()

        task_log.row_count = total_affected
        task_log.duration = round(duration, 2)

        task.last_run_at = timezone.now()
        task.last_status = 'success'
        task.save()

        if task.notify_channel != 'none':
            content = format_task_notification(task_name=task.name, success=True, row_count=total_affected, duration=duration)
            send_notification(channel=task.notify_channel, title='定时任务通知', content=content)

        task_log.save()
        messages.success(request, f'✅ 任务 "{task.name}" 执行成功，耗时 {duration:.2f}秒，影响 {total_affected} 行')
    except Exception as e:
        log.status = 'failed'
        log.error_message = str(e)
        log.save()

        task_log.status = 'failed'
        task_log.error_message = str(e)
        task_log.save()

        task.last_run_at = timezone.now()
        task.last_status = 'failed'
        task.save()

        if task.notify_channel != 'none':
            content = format_task_notification(task_name=task.name, success=False, error_message=str(e))
            send_notification(channel=task.notify_channel, title='定时任务通知', content=content)

        messages.error(request, f'❌ 任务 "{task.name}" 执行失败: {str(e)[:200]}')

    return redirect('scheduler:list')


@login_required
@require_POST
def task_send_file(request, pk):
    """执行 SQL + 生成 Excel + 发送到邮箱 + 发通知"""
    import io
    import json
    import openpyxl
    from core.utils import get_engine_from_db_conn, get_default_engine, execute_sql_statements
    from notifications.email_sender import send_email_notification

    cleanup_old_exports()

    task = get_object_or_404(ScheduledTask, pk=pk)
    script = task.script
    sql_content = script.get_sql_content()

    if not sql_content or not sql_content.strip():
        messages.error(request, '脚本内容为空，无法执行')
        return redirect('scheduler:list')

    if not task.email_recipients or not task.email_recipients.strip():
        messages.error(request, '请先配置收件人邮箱')
        return redirect('scheduler:list')

    recipients = [r.strip() for r in task.email_recipients.split(',') if r.strip()]

    task_log = ScheduledTaskLog.objects.create(
        task=task,
        task_name=task.name,
        script_name=script.name,
        sql_executed=sql_content,
        trigger_type='send_file',
        status='success',
        notify_channel='email',
        notify_target=', '.join(recipients),
        executed_by=request.user,
    )

    log = ExecutionLog.objects.create(
        script=script,
        trigger_type='manual',
        status='running',
        executed_by=request.user,
    )

    try:
        if script.target_db:
            engine = get_engine_from_db_conn(script.target_db)
        else:
            engine = get_default_engine()

        all_results, all_output, total_affected, duration = execute_sql_statements(engine, sql_content)

        log.status = 'success'
        log.output = json.dumps(all_results, ensure_ascii=False, default=str)
        log.duration = round(duration, 2)
        log.row_count = total_affected
        log.save()

        task_log.row_count = total_affected
        task_log.duration = round(duration, 2)

        # 生成 Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for idx, result in enumerate(all_results):
            if 'columns' in result and 'rows' in result:
                ws = wb.create_sheet(title=f"结果集{idx + 1}")
                ws.append(result['columns'])
                for row in result['rows']:
                    ws.append([str(v) if v is not None else '' for v in row])
            elif 'message' in result:
                ws = wb.create_sheet(title=f"信息{idx + 1}")
                ws.append([result['message']])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_bytes = output.getvalue()
        filename = f"{task.name}_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"

        from notifications.base import format_task_notification
        email_content = format_task_notification(task_name=task.name, success=True, row_count=total_affected, duration=duration)
        email_sent = send_email_notification(
            title=f'BI工具箱 - SQL执行结果: {task.name}',
            content=email_content,
            attachment_bytes=excel_bytes,
            attachment_filename=filename,
            recipients=recipients,
        )

        task_log.notify_status = email_sent

        if task.notify_channel != 'none':
            email_status = f'成功发送到 {", ".join(recipients)}' if email_sent else '发送失败（请检查邮件配置）'
            notify_content = f'**BI工具箱 - 邮件发送通知**\n\n' \
                f'| 项目 | 详情 |\n' \
                f'|------|------|\n' \
                f'| 任务名称 | {task.name} |\n' \
                f'| 发送结果 | {"✅ " + email_status if email_sent else "❌ " + email_status} |\n' \
                f'| 收件人 | {", ".join(recipients)} |\n' \
                f'| 附件文件 | {filename} |\n' \
                f'| 影响行数 | {total_affected} 行 |\n' \
                f'| 执行时间 | {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M:%S")} |'
            send_notification(channel=task.notify_channel, title='邮件发送通知', content=notify_content)

        task.last_run_at = timezone.now()
        task.last_status = 'success'
        task.save()
        task_log.save()

        if email_sent:
            messages.success(request, f'✅ 任务 "{task.name}" 执行成功，结果已发送到 {", ".join(recipients)}')
        else:
            messages.warning(request, f'✅ 任务执行成功，但邮件发送失败（请检查邮件配置）')

    except Exception as e:
        log.status = 'failed'
        log.error_message = str(e)
        log.save()

        task_log.status = 'failed'
        task_log.error_message = str(e)
        task_log.save()

        task.last_run_at = timezone.now()
        task.last_status = 'failed'
        task.save()

        if task.notify_channel != 'none':
            content = format_task_notification(task_name=task.name, success=False, error_message=str(e))
            send_notification(channel=task.notify_channel, title='定时任务通知', content=content)

        messages.error(request, f'❌ 任务 "{task.name}" 执行失败: {str(e)[:200]}')

    return redirect('scheduler:list')


def cleanup_old_exports():
    """清理超过24小时的导出文件"""
    import os
    import glob
    from datetime import datetime, timedelta
    exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'media', 'exports')
    if not os.path.exists(exports_dir):
        return
    cutoff = datetime.now() - timedelta(hours=24)
    for filepath in glob.glob(os.path.join(exports_dir, '*.xlsx')):
        try:
            file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
            if file_time < cutoff:
                os.remove(filepath)
                logger.info(f"已清理过期导出文件: {filepath}")
        except Exception as e:
            logger.warning(f"清理文件失败: {filepath}, 错误: {e}")


@login_required
def execution_logs(request):
    """执行日志列表"""
    from .models import ScheduledTaskLog
    task_logs = ScheduledTaskLog.objects.select_related('task', 'executed_by').all()[:200]
    return render(request, 'scheduler/logs.html', {'task_logs': task_logs})


@login_required
def log_detail(request, pk):
    """日志详情 API"""
    from .models import ScheduledTaskLog
    log = get_object_or_404(ScheduledTaskLog, pk=pk)
    return JsonResponse({
        'task_name': log.task_name,
        'script_name': log.script_name,
        'sql_executed': log.sql_executed,
        'trigger_type': log.get_trigger_type_display(),
        'status': log.status,
        'status_display': log.get_status_display(),
        'row_count': log.row_count,
        'duration': log.duration,
        'error_message': log.error_message,
        'notify_channel': log.notify_channel,
        'notify_target': log.notify_target,
        'notify_status': log.notify_status,
        'executed_by': str(log.executed_by) if log.executed_by else '系统',
        'executed_at': log.executed_at.strftime('%Y-%m-%d %H:%M:%S'),
    })
